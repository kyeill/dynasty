"""
Readers for manual roster extracts.

Two very different shapes arrive here:

  * a clean Fantrax league export (CSV, one row per player, carries player ids)
  * a raw copy-paste of Yahoo's roster pages into a spreadsheet (XLSX)

The Yahoo pastes are messy but regular. Every team block starts with a "Pos"
header, and the team name is the last non-empty cell above it. Yahoo's own UI
text comes along for the ride, so the trick is finding the CLEAN name rather
than the decorated one:

    NBA (single column, stacked vertically)
        G
        Nolan Traore                  <- clean
        Nolan TraoreNew Player Note   <- decorated
        BKN - G                       <- anchor
        10:00 pm @ POR                <- sometimes absent

    NFL (two columns)
        QB | Lamar Jackson                                       <- clean
           | Lamar JacksonVideo ForecastNo new player NotesBal - QB

XLSX is read with openpyxl, not pandas -- see common.py for why.
"""
from __future__ import annotations

import re
from pathlib import Path

from common import read_csv

# The position half must be upper-case codes, and the tail form additionally
# requires real spaces around the hyphen. Without both, "N. Alexander-Walker"
# ends in something that looks exactly like "TEAM - POS" ("der-Walker") and
# gets truncated to "N. Alexan".
TEAM_POS = re.compile(r"^[A-Za-z]{2,3}\s*-\s*[A-Z]{1,3}(?:,[A-Z]{1,3})*$")
TEAM_POS_TAIL = re.compile(r"[A-Za-z]{2,3}\s+-\s+[A-Z]{1,3}(?:,[A-Z]{1,3})*$")

# Yahoo welds UI text onto the name, and for names it abbreviates there is no
# clean copy anywhere -- "G. AntetokounmpoINJPlayer Note" is all you get.
NOTE_TAIL = re.compile(
    r"(?:Video Forecast|New Player Note|No new player Notes?|Player Notes?)+$")
# Requires a lower-case letter in front so a legitimately capitalised name
# ending cannot be mistaken for a status flag.
STATUS_TAIL = re.compile(r"(?<=[a-zà-ÿ])(?:INJ|GTD|DTD|OUT|SUSP|NA|IL|IR|Q|O)$")


def _clean(value) -> str:
    """Cell to trimmed text, with Yahoo's decorative characters removed."""
    if value is None:
        return ""
    s = str(value)
    if s.lower() == "nan":
        return ""
    s = s.replace("\xa0", " ")                 # non-breaking space
    s = re.sub(r"[-]", "", s)      # private-use icon glyphs
    return s.strip()


def strip_decoration(text: str) -> str:
    """'Pascal SiakamGTDNew Player Note' -> 'Pascal Siakam'.

    A denylist, and therefore a fallback only. The decoration is unbounded --
    two roster files alone contain 270 distinct suffix combinations -- and
    Yahoo can add a flag whenever it likes. Prefer known_prefix().
    """
    prev = None
    while prev != text:
        prev = text
        text = TEAM_POS_TAIL.sub("", text).strip()
        text = NOTE_TAIL.sub("", text).strip()
        text = STATUS_TAIL.sub("", text).strip()
    return text


def known_prefix(text: str, known: set) -> str | None:
    """The longest known player name that `text` starts with, or None.

    The robust way to undecorate: rather than enumerating the junk Yahoo
    appends -- which is open-ended -- ask which real player name the cell
    begins with. New injury flags and new UI strings cost nothing.

    The character after the match must not be lower-case, so "Jalen Green"
    cannot be pulled out of "Jalen Greenwood". Longest match wins.
    """
    for end in range(len(text), 2, -1):
        if text[:end] in known:
            nxt = text[end:end + 1]
            if nxt == "" or not nxt.islower():
                return text[:end]
    return None


def _sheet_rows(path) -> list[list[str]]:
    """XLSX -> grid of cleaned strings, via openpyxl."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        return [[_clean(c) for c in row] for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def _team_blocks(col0: list[str]) -> list:
    """(team_name, start, end) for each roster block, split on 'Pos' headers."""
    headers = [i for i, v in enumerate(col0) if v == "Pos"]
    blocks = []
    for n, h in enumerate(headers):
        name = next((col0[j] for j in range(h - 1, -1, -1) if col0[j]), f"Team {n + 1}")
        end = headers[n + 1] - 1 if n + 1 < len(headers) else len(col0)
        blocks.append((name, h, end))
    return blocks


def read_yahoo_paste(path, known: set | None = None) -> list[dict]:
    """Local .xlsx fallback. The live path is parse_yahoo_grid via the Sheet."""
    return parse_yahoo_grid(_sheet_rows(path), known)


def parse_yahoo_grid(grid: list[list[str]], known: set | None = None) -> list[dict]:
    """Parse pasted Yahoo roster pages into player/fantasy_team.

    Takes a grid rather than a file so the same code serves a Google Sheet tab
    (fetched as CSV) and a local .xlsx.

    `known` is the naming authority's player names. When supplied, decorated
    cells are resolved by longest-known-prefix, which is immune to new Yahoo
    suffixes. Without it the parser falls back to stripping known decoration.
    """
    known = known or set()
    grid = [[_clean(c) for c in row] for row in grid]
    if not grid:
        return []
    width = max(len(r) for r in grid)
    col0 = [(r[0] if len(r) > 0 else "") for r in grid]
    col1 = [(r[1] if len(r) > 1 else "") for r in grid] if width > 1 else [""] * len(grid)

    blocks = _team_blocks(col0)
    if not blocks:
        print("[warn] roster: no 'Pos' header rows found -- is this a Yahoo paste?")
        return []

    vertical = any(TEAM_POS.match(v) for v in col0)
    rows = []

    for team, start, end in blocks:
        if vertical:
            # Anchor on the "TEAM - POS" line; the clean name sits two above,
            # with the decorated duplicate between.
            for i in range(start, end):
                if not TEAM_POS.match(col0[i]):
                    continue
                cand = col0[i - 2] if i >= 2 else ""
                decorated = col0[i - 1] if i >= 1 else ""
                name = known_prefix(decorated, known) if decorated else None
                if name is None:
                    if cand and decorated.startswith(cand):
                        name = cand
                    elif decorated:
                        name = strip_decoration(decorated)
                    else:
                        continue
                tm, _, ps = col0[i].partition("-")
                rows.append({"name": name, "fantasy_team": team,
                             "team": tm.strip(), "pos": ps.strip()})
        else:
            # Two-column: slot position in col0, clean name in col1. The
            # decorated duplicate is the next row, with an empty col0.
            for i in range(start + 1, end):
                if not col0[i] or not col1[i]:
                    continue
                decorated = col1[i + 1] if i + 1 < len(col1) else ""
                real_team, real_pos = "", ""
                if decorated:
                    m = TEAM_POS_TAIL.search(decorated)
                    if m:
                        # "Bal - QB": the position half is what a roster append
                        # needs for a player the authority never listed.
                        real_team, _, real_pos = m.group(0).partition("-")
                        real_team = real_team.strip()
                        real_pos = real_pos.strip()
                name = (known_prefix(col1[i], known)
                        or known_prefix(decorated, known)
                        or col1[i])
                rows.append({"name": name, "fantasy_team": team,
                             "team": real_team, "pos": real_pos})

    return [r for r in rows if len(r["name"]) > 1]


def parse_mapped_grid(grid, id_col, player_col, team_col) -> list[dict]:
    """A clean export (Fantrax) whose columns are declared in config."""
    if not grid or len(grid) < 2:
        return []
    header = [_clean(h) for h in grid[0]]
    raw = [dict(zip(header, [_clean(c) for c in row])) for row in grid[1:]]
    return _mapped_rows(raw, id_col, player_col, team_col)


def read_mapped_csv(path, id_col, player_col, team_col) -> list[dict]:
    """Local CSV fallback. The live path is parse_mapped_grid via the Sheet."""
    return _mapped_rows(read_csv(path), id_col, player_col, team_col)


def _mapped_rows(raw, id_col, player_col, team_col) -> list[dict]:
    if not raw:
        return []
    missing = [c for c in (player_col, team_col) if c not in raw[0]]
    if missing:
        raise KeyError(f"missing column(s) {missing}; found {list(raw[0])}")
    out = []
    for r in raw:
        row = {"name": str(r.get(player_col, "")).strip(),
               "fantasy_team": str(r.get(team_col, "")).strip(),
               "team": str(r.get("Team", "")).strip(),
               # Read like "Team": present in a Fantrax export, absent from a
               # Yahoo paste, and needed for a player the authority never
               # heard of -- his roster row is then the ONLY place a position
               # for him exists.
               "pos": str(r.get("Position", "")).strip()}
        if id_col and id_col in r:
            # Fantrax wraps its ids in asterisks on export: *05ucd*
            row["source_id"] = str(r.get(id_col, "")).strip().strip("*")
        out.append(row)
    return out
